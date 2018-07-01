# -*- coding: utf-8 -*-
# Создаем файл для загрузки в ПАФ с неправильными, отсутствующими в Сатурне СНИЛСАми

import sys
import datetime
import time
import csv
from mysql.connector import MySQLConnection, Error
from openpyxl import Workbook
import openpyxl

from lib import read_config, format_phone, lenl, s_minus, s, l, filter_rus_sp, filter_rus_minus
########################################################################################################################
# ЗАПОЛНЕНИЕ Агент_Ид, Подписант_Ид, Пред_Страховщик_Ид
AGENT_ID = '9588'
PODPISANT_ID = '201'
PREDSTRAH_ID = '1'
########################################################################################################################
HEADS_RESULT_EXCEL_FILE = ['СНИЛС',
                          'Фамилия', 'Имя', 'Отчество',
                          'Фамилия_при_рождении', 'Имя_при_рождении', 'Отчество_при_рождении',
                          'Пол(0_мужской,1_женский)',
                          'Дата_рождения',
                          'Страна_рождения', 'Область_рождения', 'Район_рождения', 'Город_рождения',
                          'Паспорт_серия', 'Паспорт_номер', 'Паспорт_дата', 'Паспорт_Кем выдан',
                          'Паспорт_Код подразделения',

                          'Адрес_регистрации_Индекс',
                          'Адрес_регистрации_Регион', 'Адрес_регистрации_Тип_региона',
                          'Адрес_регистрации_Район', 'Адрес_регистрации_Тип_района',
                          'Адрес_регистрации_Город', 'Адрес_регистрации_Тип_города',
                          'Адрес_регистрации_Населенный_пункт', 'Адрес_регистрации_Тип_населенного_пункта',
                          'Адрес_регистрации_Улица',
                          'Адрес_регистрации_Тип_улицы',
                          'Адрес_регистрации_Дом',
                          'Адрес_регистрации_Корпус',
                          'Адрес_регистрации_Квартира',

                          'Адрес_проживания_Индекс',
                          'Адрес_проживания_Регион', 'Адрес_проживания_Тип_региона',
                          'Адрес_проживания_Район', 'Адрес_проживания_Тип_района',
                          'Адрес_проживания_Город', 'Адрес_проживания_Тип_города',
                          'Адрес_проживания_Населенный_пункт', 'Адрес_проживания_Тип_населенного_пункта',
                          'Адрес_проживания_Улица', 'Адрес_проживания_Тип_улицы',
                          'Адрес_проживания_Дом',
                          'Адрес_проживания_Корпус',
                          'Адрес_проживания_Квартира',

                          'Мобильный_телефон', 'Телефон_родственников', 'Телефон_домашний',
                          'Агент_Ид', 'Подписант_Ид', 'Пред_Страховщик_Ид'
                          ]
HEADS_SQL = ['cl.number',
            'cl.p_surname', 'cl.p_name', 'cl.p_lastname',
            'cl.b_surname', 'cl.b_name', 'cl.b_lastname',
            'cl.gender',
            'cl.b_date',
            'cl.b_country', 'cl.b_region', 'cl.b_district', 'cl.b_place',
            'cl.p_seria', 'cl.p_number', 'cl.p_date', 'cl.p_police', 'cl.p_police_code',

            'cl.p_postalcode', 'cl.p_region', 'cl.p_region_type', 'cl.p_district', 'cl.p_district_type',
            'cl.p_place', 'cl.p_place_type', 'cl.p_subplace', 'cl.p_subplace_type', 'cl.p_street',
            'cl.p_street_type', 'cl.p_building', 'cl.p_corpus', 'cl.p_flat',

            'cl.d_postalcode', 'cl.d_region', 'cl.d_region_type', 'cl.d_district', 'cl.d_district_type',
            'cl.d_place', 'cl.d_place_type', 'cl.d_subplace', 'cl.d_subplace_type', 'cl.d_street',
            'cl.d_street_type', 'cl.d_building', 'cl.d_corpus', 'cl.d_flat',

            'phone_personal_mobile', 'phone_relative_mobile', 'phone_home'
            ]
########################################################################################################################
IN_SNILS = ['СНИЛС', 'СтраховойНомер', 'Страховой_номер', 'Страховой Номер', 'Номер СНИЛС']
########################################################################################################################


def checksum(snils_dig):                         # Вычисляем 2 последних цифры СНИЛС по первым 9-ти
    def snils_csum(snils):
        k = range(9, 0, -1)
        pairs = zip(k, [int(x) for x in snils.replace('-', '').replace(' ', '')])
        return sum([k * v for k, v in pairs])

    snils = '{0:09d}'.format(snils_dig)
    csum = snils_csum(snils)
    while csum > 101:
        csum %= 101
    if csum in (100, 101):
        csum = 0

    return csum

dbconfig = read_config(filename='gen_snils.ini', section='main_mysql')
dbconn = MySQLConnection(**dbconfig)

workbooks =  []
sheets = []
for i, xlsx_file in enumerate(sys.argv):                              # Загружаем все xlsx файлы
    if i == 0:
        continue
    workbooks.append(openpyxl.load_workbook(filename=xlsx_file, read_only=True))
    sheets.append(workbooks[i-1][workbooks[i-1].sheetnames[0]])

total_rows = 0
sheets_keys = []
keys = {}
for i, sheet in enumerate(sheets):                                    # Маркируем нужные столбцы
    total_rows += sheet.max_row
    keys = {}
    for j, row in enumerate(sheet.rows):
        if j > 0:
            break
        for k, cell in enumerate(row):                                # Проверяем, чтобы был СНИЛС
            if cell.value in IN_SNILS:
                keys[IN_SNILS[0]] = k
        if len(keys) < 1:
            print('В файле "' + sys.argv[i+1] + '" отсутствует колонка со СНИЛС')
            time.sleep(3)
            sys.exit()
    sheets_keys.append(keys)

print('\n'+ datetime.datetime.now().strftime("%H:%M:%S") +' Начинаем \n')

wb = Workbook(write_only=True)
ws = wb.create_sheet('Лист1')
ws.append(HEADS_RESULT_EXCEL_FILE)                        # добавляем первую строку xlsx
wb_comp = Workbook(write_only=True)
ws_comp = wb_comp.create_sheet('Лист1')
ws_comp.append(['Реальный СНИЛС','Псевдо-СНИЛС'])                       # добавляем первую строку xlsx

dbcursor = dbconn.cursor()
dbcursor.execute('SELECT min(`number`) FROM  saturn_crm.clients WHERE `number` > 99000000000 and subdomain_id = 2;')
dbrows = dbcursor.fetchall()
start_snils = int('{0:011d}'.format(dbrows[0][0])[:-2])    # 9 цифр неправильного СНИЛСа с которого уменьшаем

sql = 'SELECT '
for i, head_sql in enumerate(HEADS_SQL):
    if i == 0:
        sql += head_sql
        continue
    sql += ', ' + head_sql
sql += ' FROM  saturn_crm.clients AS cl WHERE cl.`number` = %s'
cached_snils = 0
all_count = 0
for i, sheet in enumerate(sheets):
    for j, row in enumerate(sheet.rows):
        if j == 0:
            continue
        dbcursor = dbconn.cursor()
        dbcursor.execute(sql,(l(row[keys[IN_SNILS[0]]].value),))
        dbrows =  dbcursor.fetchall()
        count_snils = 1
        while count_snils > 0:
            start_snils -= 1
            checksum_snils = checksum(start_snils)
            cached_snils = 0
            for i in range(0, 99):
                if i != checksum_snils:
                    full_snils = start_snils * 100 + i
                    dbcursor = dbconn.cursor()
                    dbcursor.execute('SELECT `number` FROM clients WHERE `number` = %s', (full_snils,))
                    dbchk = dbcursor.fetchall()
                    if len(dbchk) == 0:
                        cached_snils = full_snils
                        count_snils -= 1
        excel_row = []
        phones = []
        for k, cell_name in enumerate(HEADS_RESULT_EXCEL_FILE):
            if k == 0:
                excel_row.append(cached_snils)
            elif k == 8 or k == 15:
                if dbrows[0][k] == None or dbrows[0][k] == 0 or dbrows[0][k] > datetime.date.today() \
                        or dbrows[0][k] < datetime.date(1868,1,1):
                    excel_row.append('1111-11-11')
                else:
                    excel_row.append(dbrows[0][k])
            elif k == 13:
                if dbrows[0][k] == None or dbrows[0][k] == 0:
                    excel_row.append('1111')
                else:
                    excel_row.append('{0:04d}'.format(dbrows[0][k]))
            elif k == 14 or k == 18 or k == 32:
                if dbrows[0][k] == None or dbrows[0][k] == 0:
                    excel_row.append('111111')
                else:
                    excel_row.append('{0:06d}'.format(dbrows[0][k]))
            elif k == 46 or k == 47 or k == 48:
                new_phone = True
                for phone in phones:
                    if phone == format_phone(dbrows[0][k]):
                        new_phone = False
                if new_phone:
                    excel_row.append(format_phone(dbrows[0][k]))
                    phones.append(format_phone(dbrows[0][k]))
                else:
                    excel_row.append(None)
            elif k == 49:
                excel_row.append(AGENT_ID)
            elif k == 50:
                excel_row.append(PODPISANT_ID)
            elif k == 51:
                excel_row.append(PREDSTRAH_ID)
            else:
                excel_row.append(dbrows[0][k])
        ws.append(excel_row)
        ws_comp.append([dbrows[0][0],cached_snils])
        all_count += 1
        if all_count % 100 == 0:
            print(datetime.datetime.now().strftime("%H:%M:%S"), 'Сконвертировано', all_count, 'СНИЛС')

#    dbcursor = dbconn.cursor()
#    dbcursor.executemany('UPDATE ' + TABLE + ' SET number_dub = %s WHERE number_dub = 0 AND (mobile0 IS NOT NULL '
#                         'OR mobile1 IS NOT NULL OR mobile2 IS NOT NULL) LIMIT 1;', cached_snils)
#    dbconn.commit()
#    print(datetime.datetime.now().strftime("%H:%M:%S"), start_snils)
wb.save(sys.argv[1][0:sys.argv[1].rfind('.xlsx')] + '_l.xlsx')
wb_comp.save(sys.argv[1][0:sys.argv[1].rfind('.xlsx')] + '_comparing.xlsx')
dbconn.close()
